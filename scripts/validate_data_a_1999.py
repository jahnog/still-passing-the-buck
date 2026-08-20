#!/usr/bin/env python3
"""Validate the committed paper-author workbook limited to 1853-1999.

The committed data/provided/data_a_1999.xlsx is the only paper-author Excel input used by:
    - scripts/generate_fiscal_fpi-fiscal.py  (cols G–J, 1853–1999)
    - scripts/generate_exchange_paper-devaluation.py  (col E, 1853–1999)
    - scripts/generate_historical_data-a-1999-excel.py  (cols D–F, 1853–1963)

Fiscal and debt data for 2000 onward are sourced from official Argentine data:
    - Debt/GDP, Debt/Exports: Secretaría de Finanzas A.2.5 + World Bank WDI
    - Result/Revenue, Result/DebtServ: datos.gob.ar SSPM dataset 379 (SPN base-caja AIF)
      downloaded by scripts/download_hacienda_spn-base-caja-historical.py
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.data_io import PROVIDED_ROOT

DEST = PROVIDED_ROOT / "data_a_1999.xlsx"


def validate_data_a_1999(path: Path) -> None:
    """Validate that Hoja1 contains paper-author rows only through 1999."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Hoja1"]

    years: list[int] = []
    for row_idx in range(1, ws.max_row + 1):
        cell_b = ws.cell(row_idx, 2).value
        try:
            years.append(int(cell_b))
        except (TypeError, ValueError):
            continue
    if not years:
        raise RuntimeError(f"No year rows found in {path}; check Hoja1 structure.")
    if min(years) != 1853 or max(years) != 1999:
        raise RuntimeError(
            f"{path} covers {min(years)}-{max(years)}; expected 1853-1999 only."
        )
    if any(year >= 2000 for year in years):
        raise RuntimeError(f"{path} contains post-1999 rows; those must come from official sources.")


def main() -> int:
    if not DEST.exists():
        raise RuntimeError(f"{DEST} is missing; the archived paper-author workbook must be committed.")
    if DEST.stat().st_size < 10_000:
        raise RuntimeError(f"{DEST} is suspiciously small.")
    validate_data_a_1999(DEST)
    print(f"Validated {DEST}: covers 1853-1999 only.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
